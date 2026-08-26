import sys
import os
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def convert_pdf_to_excel(pdf_path, xlsx_path):
    wb = openpyxl.Workbook()
    # Remove initial default sheet
    wb.remove(wb.active)

    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    header_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True)
    regular_font = Font(name="Calibri", size=10)

    with pdfplumber.open(pdf_path) as pdf:
        if len(pdf.pages) == 0:
            ws = wb.create_sheet(title="Sheet1")
            wb.save(xlsx_path)
            return 0

        for page_idx, page in enumerate(pdf.pages):
            sheet_title = f"Page {page_idx + 1}"
            ws = wb.create_sheet(title=sheet_title)
            ws.views.sheetView[0].showGridLines = True

            current_row = 1

            # 1. Extract tables
            tables = page.extract_tables()
            processed_lines = set()

            if tables:
                for table in tables:
                    if not table:
                        continue

                    # Filter out completely empty rows
                    valid_rows = [row for row in table if any(c and str(c).strip() for c in row)]
                    if not valid_rows:
                        continue

                    for r_idx, row in enumerate(valid_rows):
                        for c_idx, cell_value in enumerate(row):
                            cell = ws.cell(row=current_row, column=c_idx + 1)
                            val_str = str(cell_value or '').strip()

                            # Auto-cast numbers
                            if val_str.replace('.', '', 1).replace('-', '', 1).isdigit():
                                try:
                                    cell.value = float(val_str) if '.' in val_str else int(val_str)
                                except ValueError:
                                    cell.value = val_str
                            else:
                                cell.value = val_str

                            cell.border = thin_border
                            cell.alignment = Alignment(vertical="center", wrap_text=True)

                            if r_idx == 0:
                                cell.font = header_font
                                cell.fill = header_fill
                            else:
                                cell.font = regular_font

                            if val_str:
                                for line in val_str.split('\n'):
                                    processed_lines.add(line.strip())

                        current_row += 1

                    current_row += 1

            # 2. Extract regular text lines
            text = page.extract_text(layout=False)
            if text:
                for line in text.split('\n'):
                    cleaned_line = line.strip()
                    if not cleaned_line or cleaned_line in processed_lines:
                        continue

                    parts = [p.strip() for p in cleaned_line.split('   ') if p.strip()]
                    if not parts:
                        parts = [cleaned_line]

                    for c_idx, part in enumerate(parts):
                        cell = ws.cell(row=current_row, column=c_idx + 1)
                        cell.value = part
                        cell.font = regular_font
                        cell.alignment = Alignment(vertical="center")

                    current_row += 1

            # Set column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        for l in str(cell.value).split('\n'):
                            max_len = max(max_len, len(str(l)))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Ensure output directory exists and save
    os.makedirs(os.path.dirname(os.path.abspath(xlsx_path)), exist_ok=True)
    wb.save(xlsx_path)
    return 0

if __name__ == '__main__':
    if len(sys.argv) < 3:
        sys.stderr.write("Usage: python convert_pdf2excel.py <input.pdf> <output.xlsx>\n")
        sys.exit(1)

    in_pdf = sys.argv[1]
    out_xlsx = sys.argv[2]

    try:
        sys.exit(convert_pdf_to_excel(in_pdf, out_xlsx))
    except Exception as e:
        sys.stderr.write(f"PDF to Excel error: {str(e)}\n")
        sys.exit(1)